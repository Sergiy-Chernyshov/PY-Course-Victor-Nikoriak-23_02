from datetime import datetime, timedelta
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


SOURCE_SHEET_NAMES = ["Вибраний список", "Base"]

COL_SENDER = "Відправник"
COL_RECEIVER = "Отримувач"
COL_DEPARTMENT = "Відділ"
COL_PPT_NUMBER = "№ Ппт"
COL_PPT_DATE = "Дата Ппт"
COL_STATUS = "Статус Ппт"
COL_AMOUNT = "Сума Ппт"
COL_TYPE = "тип"
COL_MONTH = "місяць"

REQUIRED_COLUMNS = [
    COL_SENDER,
    COL_RECEIVER,
    COL_DEPARTMENT,
    COL_PPT_NUMBER,
    COL_PPT_DATE,
    COL_STATUS,
    COL_AMOUNT,
]

REPORT_TYPES = [
    "ON",
    "ON АОТ",
    "експреси",
    "торгові відділи",
]

REPORT_TITLES = {
    "ON": "Повернення товару з Онлайн магазину",
    "ON АОТ": "Товар знаходиться на АОТ онлайн магазину. Статус ППТ: В1",
    "експреси": "Недопроведені ППТ до повного циклу по Експресам",
    "торгові відділи": "Торгові відділи не прийняли товар",
}


def normalize(value):
    if pd.isna(value):
        return ""

    text = str(value).strip().upper()

    replacements = {
        "B": "В",
        "F": "Ф",
        "C": "С",
        " ": "",
        "–": "-",
        "—": "-",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return text


def is_on(value):
    return normalize(value) == "ON"


def is_dp(value):
    return normalize(value) == "DP"


def is_express(value):
    text = normalize(value)

    return bool(
        re.match(r"^X\d+", text)
        or re.match(r"^I\d+", text)
        or re.match(r"^І\d+", text)
    )


def get_month_from_date(value):
    if pd.isna(value):
        return ""

    if isinstance(value, datetime):
        return value.month

    if isinstance(value, (int, float)):
        excel_start = datetime(1899, 12, 30)
        return (excel_start + timedelta(days=int(value))).month

    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)

    if pd.isna(parsed):
        return ""

    return parsed.month


def is_trade_status(status):
    status = normalize(status)

    trade_statuses = {
        "В1",
        "Ф1",
        "Ф2",
        "С1",
        "С2",
        "В1-Ф1",
        "В1-Ф1-Ф2",
        "В1-Ф1-Ф2-С1-С2",
    }

    return status in trade_statuses


def is_empty_status(status):
    return normalize(status) == ""


def detect_type(row):
    sender = row.get(COL_SENDER, "")
    receiver = row.get(COL_RECEIVER, "")
    status = row.get(COL_STATUS, "")

    sender_is_on = is_on(sender)
    receiver_is_on = is_on(receiver)

    sender_is_dp = is_dp(sender)
    receiver_is_dp = is_dp(receiver)

    sender_is_express = is_express(sender)
    receiver_is_express = is_express(receiver)

    status_text = normalize(status)

    if sender_is_dp and receiver_is_on and status_text == "В1":
        return "ON АОТ"

    if sender_is_express or receiver_is_express:
        return "експреси"

    if receiver_is_dp and is_trade_status(status):
        return "торгові відділи"

    if sender_is_dp and (is_empty_status(status) or is_trade_status(status)):
        return "торгові відділи"

    if (sender_is_on or receiver_is_on) and not is_trade_status(status):
        return "ON"

    return "видалити"


def make_report_row(label, dataframe, months):
    row = {
        "label": label,
        "months": [],
        "total_count": 0,
        "total_sum": 0,
    }

    for month in months:
        month_df = dataframe[dataframe[COL_MONTH] == month]

        count = int(month_df[COL_PPT_NUMBER].count())
        total = round(float(month_df[COL_AMOUNT].sum()), 2)

        row["months"].append(
            {
                "month": month,
                "count": count,
                "sum": total,
            }
        )

        row["total_count"] += count
        row["total_sum"] += total

    row["total_sum"] = round(row["total_sum"], 2)

    return row


def build_report_blocks(dataframe):
    months = sorted(
        int(month)
        for month in dataframe[COL_MONTH].dropna().unique()
        if str(month).strip() != ""
    )

    blocks = []

    for report_type in REPORT_TYPES:
        type_df = dataframe[dataframe[COL_TYPE] == report_type]

        rows = [make_report_row(report_type, type_df, months)]

        departments = type_df[COL_DEPARTMENT].dropna().sort_values().unique()

        for department in departments:
            department_df = type_df[type_df[COL_DEPARTMENT] == department]
            rows.append(make_report_row(str(department), department_df, months))

        blocks.append(
            {
                "type": report_type,
                "title": REPORT_TITLES[report_type],
                "rows": rows,
            }
        )

    return blocks, months


def find_source_sheet(workbook):
    for sheet_name in SOURCE_SHEET_NAMES:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

    return workbook[workbook.sheetnames[0]]


def get_headers(sheet):
    headers = {}

    for cell in sheet[3]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    return headers


def ensure_column(sheet, headers, column_name):
    if column_name in headers:
        return headers[column_name]

    new_column = sheet.max_column + 1
    sheet.cell(row=3, column=new_column).value = column_name
    headers[column_name] = new_column

    return new_column


def get_month_name(month):
    month_names = {
        1: "Січень",
        2: "Лютий",
        3: "Березень",
        4: "Квітень",
        5: "Травень",
        6: "Червень",
        7: "Липень",
        8: "Серпень",
        9: "Вересень",
        10: "Жовтень",
        11: "Листопад",
        12: "Грудень",
    }

    return month_names.get(month, f"Місяць {month}")


def get_period_title(months):
    if not months:
        return "Не проведені ППТ за період"

    first_month = get_month_name(min(months))
    last_month = get_month_name(max(months))

    if first_month == last_month:
        return f"Не проведені ППТ за період {first_month}"

    return f"Не проведені ППТ за період {first_month} - {last_month}"


def write_svod_sheet(sheet, blocks, months):
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=14)
    block_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    small_font = Font(size=8)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    total_columns = 1 + len(months) * 2 + 2

    sheet.cell(row=1, column=1).value = get_period_title(months)
    sheet.cell(row=1, column=1).font = title_font
    sheet.cell(row=1, column=1).alignment = center
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

    sheet.cell(row=2, column=1).value = (
        f"Сформовано автоматично: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    sheet.cell(row=2, column=1).font = small_font
    sheet.cell(row=2, column=1).alignment = left
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)

    current_row = 4

    for block in blocks:
        sheet.cell(row=current_row, column=1).value = block["title"]
        sheet.cell(row=current_row, column=1).font = block_font
        sheet.cell(row=current_row, column=1).alignment = left
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=total_columns,
        )
        current_row += 1

        sheet.cell(row=current_row, column=1).value = ""
        sheet.cell(row=current_row + 1, column=1).value = "ТИП"

        column = 2

        for month in months:
            sheet.cell(row=current_row, column=column).value = get_month_name(month)
            sheet.merge_cells(
                start_row=current_row,
                start_column=column,
                end_row=current_row,
                end_column=column + 1,
            )

            sheet.cell(row=current_row + 1, column=column).value = "кіл ППТ"
            sheet.cell(row=current_row + 1, column=column + 1).value = "сума ППТ"

            column += 2

        sheet.cell(row=current_row, column=column).value = "Загалом"
        sheet.merge_cells(
            start_row=current_row,
            start_column=column,
            end_row=current_row,
            end_column=column + 1,
        )

        sheet.cell(row=current_row + 1, column=column).value = "кіл ППТ"
        sheet.cell(row=current_row + 1, column=column + 1).value = "сума ППТ"

        for row_number in [current_row, current_row + 1]:
            for col_number in range(1, total_columns + 1):
                cell = sheet.cell(row=row_number, column=col_number)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center

        current_row += 2

        for report_row in block["rows"]:
            sheet.cell(row=current_row, column=1).value = report_row["label"]

            column = 2

            for month_data in report_row["months"]:
                count_cell = sheet.cell(row=current_row, column=column)
                sum_cell = sheet.cell(row=current_row, column=column + 1)

                count_cell.value = month_data["count"]
                sum_cell.value = month_data["sum"]

                count_cell.number_format = "0"
                sum_cell.number_format = "0"

                column += 2

            total_count_cell = sheet.cell(row=current_row, column=column)
            total_sum_cell = sheet.cell(row=current_row, column=column + 1)

            total_count_cell.value = report_row["total_count"]
            total_sum_cell.value = report_row["total_sum"]

            total_count_cell.number_format = "0"
            total_sum_cell.number_format = "0"

            for col_number in range(1, total_columns + 1):
                cell = sheet.cell(row=current_row, column=col_number)
                cell.border = border
                cell.alignment = center

                if col_number == 1:
                    cell.alignment = left

                if report_row["label"] == block["type"]:
                    cell.font = header_font
                    cell.fill = total_fill

            current_row += 1

        current_row += 2

    last_row = current_row - 1

    sheet.column_dimensions["A"].width = 34

    for column_number in range(2, total_columns + 1):
        column_letter = get_column_letter(column_number)
        sheet.column_dimensions[column_letter].width = 13

    sheet.freeze_panes = "B6"

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    sheet.page_margins = PageMargins(
        left=0.3,
        right=0.3,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )

    last_column_letter = get_column_letter(total_columns)
    sheet.print_area = f"A1:{last_column_letter}{last_row}"


def update_excel_file(file_path, dataframe, blocks, months):
    workbook = load_workbook(file_path)

    base_sheet = find_source_sheet(workbook)
    base_sheet.title = "Base"

    headers = get_headers(base_sheet)

    type_column = ensure_column(base_sheet, headers, COL_TYPE)
    month_column = ensure_column(base_sheet, headers, COL_MONTH)

    for index, row in dataframe.iterrows():
        excel_row = index + 4

        month_value = row[COL_MONTH]

        if pd.notna(month_value) and month_value != "":
            month_value = int(month_value)

        base_sheet.cell(row=excel_row, column=type_column).value = row[COL_TYPE]
        base_sheet.cell(row=excel_row, column=month_column).value = month_value
        base_sheet.cell(row=excel_row, column=month_column).number_format = "0"

    if "Svod" in workbook.sheetnames:
        del workbook["Svod"]

    svod_sheet = workbook.create_sheet("Svod")
    write_svod_sheet(svod_sheet, blocks, months)

    output_path = Path(file_path).with_name(f"processed_{Path(file_path).name}")
    workbook.save(output_path)

    return output_path


def create_pdf_copy(xlsx_path):
    try:
        import win32com.client
    except ImportError:
        return None

    pdf_path = Path(xlsx_path).with_suffix(".pdf")

    excel = None
    workbook = None

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(str(Path(xlsx_path).resolve()))
        worksheet = workbook.Worksheets("Svod")
        worksheet.ExportAsFixedFormat(0, str(pdf_path.resolve()))

        workbook.Close(False)

        return pdf_path

    except Exception:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass

        return None

    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass

def process_excel_file(file_path):
    workbook = load_workbook(file_path, read_only=True)
    source_sheet = find_source_sheet(workbook)
    source_sheet_name = source_sheet.title
    workbook.close()

    df = pd.read_excel(
        file_path,
        sheet_name=source_sheet_name,
        header=2,
        engine="openpyxl",
    )

    df = df.dropna(how="all")

    missing_columns = [
        column for column in REQUIRED_COLUMNS if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            "У файлі відсутні колонки: " + ", ".join(missing_columns)
        )

    df[COL_AMOUNT] = pd.to_numeric(df[COL_AMOUNT], errors="coerce").fillna(0)

    df[COL_MONTH] = df[COL_PPT_DATE].apply(get_month_from_date)
    df[COL_TYPE] = df.apply(detect_type, axis=1)

    filtered_df = df[df[COL_TYPE] != "видалити"].copy()

    blocks, months = build_report_blocks(filtered_df)
    output_path = update_excel_file(file_path, df, blocks, months)
    pdf_path = create_pdf_copy(output_path)

    totals = {
        "rows_total": len(df),
        "rows_used": len(filtered_df),
        "rows_deleted": len(df) - len(filtered_df),
        "total_count": int(filtered_df[COL_PPT_NUMBER].count()),
        "total_sum": round(float(filtered_df[COL_AMOUNT].sum()), 2),
    }

    return {
        "blocks": blocks,
        "months": months,
        "totals": totals,
        "download_url": "/media/uploads/" + output_path.name,
        "pdf_url": "/media/uploads/" + pdf_path.name if pdf_path else "",
    }


def process_internet_orders_file(file_path):
    excel_file = pd.ExcelFile(file_path, engine="openpyxl")
    sheet_name = excel_file.sheet_names[0]

    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        engine="openpyxl",
    )

    df = df.dropna(how="all")

    return {
        "sheet_name": sheet_name,
        "rows_total": len(df),
        "columns_total": len(df.columns),
        "columns": list(df.columns),
        "preview": df.head(10).fillna("").to_dict("records"),
    }